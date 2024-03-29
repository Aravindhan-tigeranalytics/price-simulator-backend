from numpy import save
import xlsxwriter
import datetime
import openpyxl
import decimal
# from openpyxl import worksheet
from datetime import datetime,timedelta
from core import models as model
from utils import models
from . import constants as const
from utils.models import ScenarioPlannerMetricModel,PromoMeta
from utils import util , roi

def _get_sheet_value(sheet , row , column):
    val = sheet.cell(row = row, column = column).value
    # print(val , "value returuning")
    return sheet.cell(row = row, column = column).value
def excel(data , output):
    ROW_CONST = 5
    COL_CONST = 1
#    data = da
#     [ {'product_group': 'ORBIT OTC', 'retailer': 'Magnit', 'category': 'Gum', 'product_group_retailer': 'ORBIT OTC', 'current_lpi': 18.02, 'increased_lpi': 18.02, 'current_rsp': 21.33, 'increased_rsp': 21.33, 'current_cogs': 4.881618, 'increased_cogs': 4.881618, 'lpi_increase': 0, 'rsp_increase': 0, 'cogs_increase': 0, 'base_price_elasticity': '-1.98', 'base_price_elasticity_manual': '-1.98', 'net_elasticity': '-1.53', 'competition': 'Not Follows', 'tonnes': 0, 'rsv': 0, 'mac': 0, 'rp': 0, 'nsv': 0, 'te': 0}, {'product_group': 'ORBIT XXL', 'retailer': 'Magnit', 'category': 'Gum', 'product_group_retailer': 'ORBIT XXL', 'current_lpi': 23.74, 'increased_lpi': 23.74, 'current_rsp': 27.44, 'increased_rsp': 27.44, 'current_cogs': 7.197967999999996, 'increased_cogs': 7.197967999999996, 'lpi_increase': 0, 'rsp_increase': 0, 'cogs_increase': 0, 'base_price_elasticity': '-1.08', 'base_price_elasticity_manual': '-1.08', 'net_elasticity': '-1.08', 'competition': 'Not Follows', 'tonnes': 0, 'rsv': 0, 'mac': 0, 'rp': 0, 'nsv': 0, 'te': 0}, {'product_group': 'BIG BARS', 'retailer': 'Magnit', 'category': 'Choco', 'product_group_retailer': 'BIG BARS', 'current_lpi': 32.46, 'increased_lpi': 32.46, 'current_rsp': 39.19, 'increased_rsp': 39.19, 'current_cogs': 12.097842, 'increased_cogs': 12.097842, 'lpi_increase': 0, 'rsp_increase': 0, 'cogs_increase': 0, 'base_price_elasticity': '-1.17', 'base_price_elasticity_manual': '-1.17', 'net_elasticity': '-0.45', 'competition': 'Not Follows', 'tonnes': 0, 'rsv': 0, 'mac': 0, 'rp': 0, 'nsv': 0, 'te': 0}, {'product_group': 'BIG BARS_SNICKERSCRISPER', 'retailer': 'Magnit', 'category': 'Choco', 'product_group_retailer': 'BIG BARS_SNICKERSCRISPER', 'current_lpi': 21.73, 'increased_lpi': 21.73, 'current_rsp': 28.32, 'increased_rsp': 28.32, 'current_cogs': 11.403904, 'increased_cogs': 11.403904, 'lpi_increase': 0, 'rsp_increase': 0, 'cogs_increase': 0, 'base_price_elasticity': '-1.19', 'base_price_elasticity_manual': '-1.19', 'net_elasticity': '-1.21', 'competition': 'Not Follows', 'tonnes': 0, 'rsv': 0, 'mac': 0, 'rp': 0, 'nsv': 0, 'te': 0}, {'product_group': 'BIG BARS_SNICKERSCRISPER_DUO', 'retailer': 'Magnit', 'category': 'Choco', 'product_group_retailer': 'BIG BARS_SNICKERSCRISPER_DUO', 'current_lpi': 17.68, 'increased_lpi': 17.68, 'current_rsp': 23.52, 'increased_rsp': 23.52, 'current_cogs': 7.667816, 'increased_cogs': 7.667816, 'lpi_increase': 0, 'rsp_increase': 0, 'cogs_increase': 0, 'base_price_elasticity': '-1.67', 'base_price_elasticity_manual': '-1.67', 'net_elasticity': '-1.67', 'competition': 'Not Follows', 'tonnes': 0, 'rsv': 0, 'mac': 0, 'rp': 0, 'nsv': 0, 'te': 0}, {'product_group': 'STD BARS', 'retailer': 'Magnit', 'category': 'Choco', 'product_group_retailer': 'STD BARS', 'current_lpi': 19.89, 'increased_lpi': 19.89, 'current_rsp': 23.8, 'increased_rsp': 23.8, 'current_cogs': 6.969456000000001, 'increased_cogs': 6.969456000000001, 'lpi_increase': 0, 'rsp_increase': 0, 'cogs_increase': 0, 'base_price_elasticity': '-1.80', 'base_price_elasticity_manual': '-1.80', 'net_elasticity': '-1.02', 'competition': 'Not Follows', 'tonnes': 0, 'rsv': 0, 'mac': 0, 'rp': 0, 'nsv': 0, 'te': 0}, {'product_group': 'STD BARS_MILKYWAY', 'retailer': 'Magnit', 'category': 'Choco', 'product_group_retailer': 'STD BARS_MILKYWAY', 'current_lpi': 11.19, 'increased_lpi': 11.19, 'current_rsp': 14.76, 'increased_rsp': 14.76, 'current_cogs': 3.6098939999999997, 'increased_cogs': 3.6098939999999997, 'lpi_increase': 0, 'rsp_increase': 0, 'cogs_increase': 0, 'base_price_elasticity': '-1.28', 'base_price_elasticity_manual': '-1.28', 'net_elasticity': '-0.35', 'competition': 'Not Follows', 'tonnes': 0, 'rsv': 0, 'mac': 0, 'rp': 0, 'nsv': 0, 'te': 0}]
    keys = list(data[0].keys())
    values = []
    for d in data:
        values.append(list(d.values()))
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    worksheet.hide_gridlines(2)
    merge_format_date = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': 'yellow'})

    merge_format_app = workbook.add_format({
        'bold': 1,
        
        'align': 'center',
        'valign': 'vcenter'
        })
    merge_format_app.set_font_size(20)
    worksheet.set_column('B:D', 12)
    format2 = workbook.add_format({'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': 'blue'})

    format1 = workbook.add_format({'bold': 0,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': 1})
    money = workbook.add_format({'num_format': '₽#,##0'})
    # worksheet.write('A2', 'Item', format2)
    # worksheet.write('B2', 'Cost', format2)
    row = ROW_CONST
    col = COL_CONST

    worksheet.merge_range('B2:D2', 'Downloaded on ' +dateformat() , merge_format_date)
    worksheet.merge_range('B3:E3', 'Price Simulator Tool' , merge_format_app)

    # Iterate over the data and write it out row by row.
    for header in keys:
        worksheet.write(row, col,     " ".join(header.split("_")),format2)
        worksheet.set_row(row, 25)
        worksheet.set_column(col,col, 20)
        col+=1
    row+=1
    col =COL_CONST
    for val in values :
        col = COL_CONST
        for v in val:
            worksheet.write(row, col, " ".join(v.split("_")),format1)
            worksheet.set_row(row, 45)
            worksheet.set_column(col,col, 20)
            col+=1
        row+=1



    # for item, cost in (expenses):
    #     worksheet.write(row, col,     item,format2)
    #     worksheet.write(row, col + 1, cost , format2)
    #     row += 1

    # worksheet.write(row, 0, 'Total' , bold)
    # worksheet.write(row, 1, '=SUM(B1:B4)' , money)

    workbook.close()
def excel_summary(data , output):
    ROW_CONST = 6
    COL_CONST = 4
    # data = {'sc1sdf': {'name': 'sc1sdf', 'header': ['units', 'tonnes', 'lsv', 'rsv', 'nsv', 'cogs', 'nsv_tonnes', 'te', 'te_percent_lsv', 'te_units', 'mac', 'mac_percent_nsv', 'rp', 'rp_percent_rsv'], 'current': ['136.0M ₽', '2.0K ₽', '2.6B ₽', '3.0B ₽', '1.6B ₽', '709.4M ₽', '815.1K ₽', '968.1M ₽', '37.5 ₽', '7.1 ₽', '907.1M ₽', '56.1 ₽', '1.4B ₽', '45.7 ₽'], 'simulated': ['123.4M ₽', '1.8K ₽', '2.7B ₽', '2.8B ₽', '1.7B ₽', '659.0M ₽', '928.8K ₽', '1.0B ₽', '37.5 ₽', '8.1 ₽', '1.0B ₽', '60.7 ₽', '1.2B ₽', '41.0 ₽'], 'Absolute change': ['-12.6M ₽', '-0.2K ₽', '95.6M ₽', '-0.1B ₽', '59.9M ₽', '-50.4M ₽', '-0.3M ₽', '35.7M ₽', '-0.0 ₽', '-2.8 ₽', '110.3M ₽', '4.6 ₽', '-0.2B ₽', '-4.7 ₽'], 'percent change': ['-9.2 %', '-9.0 %', '3.7 %', '-4.5 %', '3.7 %', '-7.1 %', '-0.4 %', '3.7 %', '-0.0 %', '-0.4 %', '12.2 %', '8.2 %', '-14.3 %', '-10.2 %']}, 'otc&xxl': {'name': 'otc&xxl', 'header': ['units', 'tonnes', 'lsv', 'rsv', 'nsv', 'cogs', 'nsv_tonnes', 'te', 'te_percent_lsv', 'te_units', 'mac', 'mac_percent_nsv', 'rp', 'rp_percent_rsv'], 'current': ['266.1M ₽', '10.2K ₽', '5.8B ₽', '6.9B ₽', '4.2B ₽', '1.9B ₽', '406.8K ₽', '1.6B ₽', '28.2 ₽', '6.1 ₽', '2.3B ₽', '55.0 ₽', '2.8B ₽', '40.1 ₽'], 'simulated': ['241.0M ₽', '9.9K ₽', '5.5B ₽', '6.6B ₽', '4.0B ₽', '1.8B ₽', '405.1K ₽', '1.5B ₽', '27.8 ₽', '6.4 ₽', '2.2B ₽', '55.0 ₽', '2.6B ₽', '39.9 ₽'], 'Absolute change': ['-25.1M ₽', '-0.4K ₽', '-0.3B ₽', '-0.3B ₽', '-0.2B ₽', '-69.5M ₽', '453.9K ₽', '-96.0M ₽', '-0.4 ₽', '3.8 ₽', '-92.3M ₽', '-0.1 ₽', '-0.1B ₽', '-0.2 ₽'], 'percent change': ['-9.4 %', '-3.5 %', '-4.4 %', '-4.3 %', '-3.9 %', '-3.7 %', '1.1 %', '-5.9 %', '-1.5 %', '0.6 %', '-4.0 %', '-0.1 %', '-4.8 %', '-0.6 %']}}
    workbook = xlsxwriter.Workbook(output)
    merge_format_date = workbook.add_format({
        'bold': 1,
        'align': 'center',
        'valign': 'vcenter',
         })

    merge_format_app = workbook.add_format({
        'bold': 1,
        
        'align': 'center',
        'valign': 'vcenter'
        })
    merge_format_app.set_font_size(20)
    worksheet = workbook.add_worksheet()
    worksheet.hide_gridlines(2)
    format_header = workbook.add_format({'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'})
    format_header.set_font_size(14)
    format_name = workbook.add_format({'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'})
    format_name.set_font_size(20)
    format_value = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'})
    # format_name.set_font_size(20)

    row = ROW_CONST
    col = COL_CONST
    
    worksheet.merge_range('B2:D2', 'Downloaded on ' +dateformat() , merge_format_date)
    worksheet.merge_range('B3:D3', 'Price Simulator Tool' , merge_format_app)
    worksheet.merge_range('B4:D4', 'Comparison Summary',merge_format_app)
    # for header in keys:
    #     worksheet.write(row, col,     header,format2)
    #     worksheet.set_row(row, 20)
    #     worksheet.set_column(col,col, 15)
    #     col+=1
    zip_list = []
    data_val = list(data.values())
    for i in data_val:
        # print(i['name'] , "i vaue")
        zip_list.append(zip(i["header"],i["current"],i["simulated"],i["Absolute change"],i["percent change"]))
    for idx,z in enumerate(zip_list):
        worksheet.merge_range(row+1,col-4,row+4,col-2 ,data_val[idx]['name'] ,format_name)
        _writeExcel(worksheet,row+1, col-1,"Current Value",format_header)
        _writeExcel(worksheet,row+2, col-1,"Simulated Value",format_header)
        _writeExcel(worksheet,row+3, col-1,"Absolute Change",format_header)
        _writeExcel(worksheet,row+4, col-1,"Percent Change",format_header)
        for header, current , simulate,absc , per in z:
            _writeExcel(worksheet,row, col," ".join(header.split("_")).title(),format_header)
            _writeExcel(worksheet,row+1, col,current,format_value)
            _writeExcel(worksheet,row+2, col,simulate,format_value)
            _writeExcel(worksheet,row+3, col,absc,format_value)
            _writeExcel(worksheet,row+4, col,per,format_value)
            col+=1  
        row+=6
        col = COL_CONST
    
    
    
    workbook.close()

def _writeExcel(worksheet , row , col , val , _format):
    worksheet.write(row, col,val,_format)
    worksheet.set_row(row, 30)
    worksheet.set_column(col,col, 20)


def dateformat():

    x = datetime.datetime.now()
    return x.strftime("%b %d %Y %H:%M:%S")
    
def read_promo_coeff(file):
    headers = const.COEFF_HEADER
    book = openpyxl.load_workbook(file,data_only=True)
    sheet = book['MODEL_COEFFICIENT']
    columns = sheet.max_column
    rows = sheet.max_row
    print(columns , rows , "columns and rows")
    col_ = [i for i in range(1,len(headers)+1)]
    row_ =0
    for row in range(row_+2 , rows+1):
        db_meta = model.ModelMeta()
        db_coeff = model.ModelCoefficient()
        for c in range(0,len(col_)):
            
            # obj = {}
            # print(row , c," :rowcvalue")'Account Name' , 'Corporate Segment' , 'PPG'
            cell_obj = sheet.cell(row = row,column = col_[c])
            if(headers[c] in const.PROMO_MODEL_META_MAP):
                print(headers[c] , const.PROMO_MODEL_META_MAP[headers[c]],cell_obj.value , "generated value")
                setattr(db_meta,const.PROMO_MODEL_META_MAP[headers[c]],cell_obj.value)
            elif(headers[c] in const.PROMO_MODEL_COEFF_MAP):
                setattr(db_coeff,const.PROMO_MODEL_COEFF_MAP[headers[c]],
                        cell_obj.value if cell_obj.value else 0.0)
            
        # print(db_meta.product_group , db_meta.corporate_segment , db_meta.account_name , "from object before save")
        db_meta.slug = util.generate_slug_string(db_meta.account_name,db_meta.corporate_segment,db_meta.product_group)
        if not model.ModelMeta.objects.filter(slug=db_meta.slug).exists():
        # import pdb
        # pdb.set_trace()
            db_meta.save()
        
        # print(db_meta.product_group , db_meta.corporate_segment , db_meta.account_name , "from object")
        # print(saved_meta , "saved meta value")
        db_coeff.model_meta = model.ModelMeta.objects.filter(
            account_name=db_meta.account_name,corporate_segment=db_meta.corporate_segment,
            product_group=db_meta.product_group).first()
        db_coeff.save()
            # ob.append(cell_obj.value)
    # print(ob , "OBBB")
            # _genObj(obj,cell_obj.value,headers[c])
    #     ob.append(ScenarioPlannerMetricModel(obj))
    # print(len(ob) , "OBJECY LIST")
    # _update_date(ob)
    book.close()


def read_promo_data(file):
    headers = const.DATA_HEADER
    book = openpyxl.load_workbook(file,data_only=True)
    sheet = book['MODEL_DATA']
    columns = sheet.max_column
    rows = sheet.max_row

    col_ = [i for i in range(1,len(headers)+1)]
    row_ =0
    for row in range(row_+2 , rows+1):
        
        db_meta = PromoMeta()
        db_data = model.ModelData()
        for c in range(0,len(col_)):
            cell_obj = sheet.cell(row = row,column = col_[c])
            if(headers[c] in const.PROMO_MODEL_META_MAP):
                setattr(db_meta,const.PROMO_MODEL_META_MAP[headers[c]],cell_obj.value)
            elif(headers[c] in const.PROMO_MODEL_DATA_MAP):
                setattr(db_data,const.PROMO_MODEL_DATA_MAP[headers[c]],cell_obj.value)
                
        # model.ModelData.objects.get(
        #     account_name=db_meta.account_name,corporate_segment=db_meta.corporate_segment,
        #     product_group=db_meta.product_group
        # )
        db_data.model_meta = model.ModelMeta.objects.get(
            
            slug=util.generate_slug_string(db_meta.account_name,
                                           db_meta.corporate_segment,
                                           db_meta.product_group)
            
        )
        db_data.save()
    book.close()
    
def read_roi_data(file):
    # mode , cre = model.ModelMeta.objects.get_or_create(
    #         account_name = 'acc',
    #         corporate_segment ='ss',
    #         product_group = 'sss')
    # import pdb
    # pdb.set_trace()
    headers = const.ROI_HEADER
    book = openpyxl.load_workbook(file,data_only=True)
    sheet = book['ROI_Data_All_retailers_with_ext']
    columns = sheet.max_column
    rows = sheet.max_row
    col_= []
    row_ =1
    header_found = False
    for row in range(1,rows+1):
        print(row , "row count")
        for col in range(1,columns+1):
            print(col , "column count")
            cell_obj = sheet.cell(row = row, column = col)
            if cell_obj.value in headers:
                header_found = True
                print(cell_obj.value , 'object value')
                # col_taken = True
                col_.append(col)
        if header_found:
           break 
    print(col_ , "coldddd")
    # import pdb
    # pdb.set_trace()
    for row in range(row_+1 , rows+1):
        
        meta , created = model.ModelMeta.objects.get_or_create(
            account_name = _get_sheet_value(sheet ,row , 1),
            corporate_segment = _get_sheet_value(sheet ,row , 2),
            product_group = _get_sheet_value(sheet ,row , 3),
            # brand_filter = _get_sheet_value(sheet ,row , 4),
            # brand_format_filter = _get_sheet_value(sheet ,row , 5),
            # strategic_cell_filter = _get_sheet_value(sheet ,row , 5),
            slug = util.generate_slug_string(
                _get_sheet_value(sheet ,row , 1),
               _get_sheet_value(sheet ,row , 2),
                _get_sheet_value(sheet ,row , 3)
            )
        )
        if created:
            # print(created , "created")
            # print(meta , "meta")
            meta.brand_filter = _get_sheet_value(sheet ,row , 4)
            meta.brand_format_filter =_get_sheet_value(sheet ,row , 5)
            meta.strategic_cell_filter =_get_sheet_value(sheet ,row , 6)
            meta.save()
            # meta.slug = util.generate_slug_string(
            #     _get_sheet_value(sheet ,row , 1),
            #    _get_sheet_value(sheet ,row , 2),
            #     _get_sheet_value(sheet ,row , 3))
        roi = model.ModelROI(
            model_meta = meta,
            year = _get_sheet_value(sheet ,row , 9),
            week = _get_sheet_value(sheet ,row , 12),
            on_inv = _get_sheet_value(sheet ,row , 17),
            off_inv = _get_sheet_value(sheet ,row , 18),
             gmac = _get_sheet_value(sheet ,row , 20),
            list_price = _get_sheet_value(sheet ,row , 23),
        )
        roi.save() 
                
    # print(col_ , "final col_")

def read_excel(loc):
    # ScenarioPlannerMetrics.objects.all().delete()
    headers = ['Category' , 'Product Group' , 'Retailer' ,'Brand Filter','Brand Format Filter',
    'Strategic Cell Filter','Year' , 'Date','Base Price Elasticity','Cross Elasticity',
    'Net Elasticity','Base Units','List Price','Retailer Median Base Price',
    'Retailer Median Base Price  w\o VAT','On Inv. %','Off Inv. %','TPR %','GMAC%, LSV',
    'Product Group Weight (grams)']
    taken_header = []
    book = openpyxl.load_workbook(loc,data_only=True)
    shet = book['Scenario Planner']
    columns = shet.max_column
    rows = shet.max_row
    col_ = []
    row_ =0
    col_taken = False
    for row in range(1,rows+1):
        for col in range(1,columns+1):
            cell_obj = shet.cell(row = row, column = col)
            if cell_obj.value in headers and cell_obj.value not in taken_header:
                col_taken = True
                col_.append(col)
                taken_header.append(cell_obj.value)
        if col_taken:
            row_ = row
            break
    obj = {}
    ob=[]
    for row in range(row_+1 , rows+1):
        obj = {}
        # metric = ScenarioPlannerMetrics()
        for c in range(0,len(col_)):
            cell_obj = shet.cell(row = row, column = col_[c])
            _genObj(obj,cell_obj.value,headers[c])
        ob.append(ScenarioPlannerMetricModel(obj))
    print(len(ob) , "OBJECY LIST")
    _update_date(ob)
        # metric.save()
    # for row in range(row_+1 , rows+1):
    #     metric = ScenarioPlannerMetrics()
    #     for c in range(0,len(col_)):
    #         cell_obj = shet.cell(row = row, column = col_[c])
    #         _updateMetric(metric , cell_obj.value,headers[c])
    #     metric.save()
    # for row in range(row_+1 , rows+1):
    #     metric = ScenarioPlannerMetrics()
    #     for c in range(0,len(col_)):
    #         cell_obj = shet.cell(row = row, column = col_[c])
    #         _updateMetricDup(metric , cell_obj.value,headers[c])
    #     metric.save()
def _update_date(obj):
    # for o in obj:
    #     metric = ScenarioPlannerMetrics()
    #     _updateMetricFromObject(metric , o)
    #     metric.save()
    # print(min(obj,key=lambda x:x.date).date , "min date")
    # print(max(obj,key=lambda x:x.date).date , "max date")
    # print(max(obj,key=lambda x:x.date).date + timedelta(days=7) , "initial date")

    li = util.grouping(obj , max(obj,key=lambda x:x.date).date + timedelta(days=7))
    for o in li:
        metric = model.ScenarioPlannerMetrics()
        _updateMetricFromObject(metric , o)
        metric.save()
    print(min(li,key=lambda x:x.date).date , "min date")
    print(max(li,key=lambda x:x.date).date , "max date")
    print(max(li,key=lambda x:x.date).date + timedelta(days=7) , "initial date")
    
def _updateMetricFromObject(metric:model.ScenarioPlannerMetrics , obj : ScenarioPlannerMetricModel):
    # print(obj.year , "object year")
     
    metric.category = obj.category
    metric.product_group = obj.product_group
    metric.retailer = obj.retailer
    metric.brand_filter = obj.brand_filter
    metric.brand_format_filter = obj.brand_format_filter
    metric.strategic_cell_filter = obj.brand_format_filter
    metric.year = obj.year
    metric.date = obj.date
    metric.base_price_elasticity = obj.base_price_elasticity
    metric.cross_elasticity = obj.cross_elasticity
    metric.net_elasticity = obj.net_elasticity
    metric.base_units = obj.base_units
    metric.list_price = obj.list_price
    metric.retailer_median_base_price = obj.retailer_median_base_price
    metric.retailer_median_base_price_w_o_vat = obj.retailer_median_base_price_w_o_vat
    metric.on_inv_percent = obj.on_inv_percent
    metric.off_inv_percent = obj.off_inv_percent
    metric.tpr_percent = obj.tpr_percent 
    metric.gmac_percent_lsv = obj.gmac_percent_lsv
    metric.product_group_weight = obj.product_group_weight

def _updateMetric(metric:model.ScenarioPlannerMetrics , value,header):
    if(header == 'Category'):
        metric.category = value.strip()
    elif(header == 'Product Group'):
        metric.product_group = value.strip()
    elif(header == 'Retailer'):
        metric.retailer = value.strip()
    elif(header == 'Brand Filter'):
        metric.brand_filter = value.strip()
    elif(header == 'Brand Format Filter'):
        metric.brand_format_filter = value.strip()
    elif(header == 'Strategic Cell Filter'):
        metric.strategic_cell_filter = value.strip()
    elif(header == 'Year'):
        metric.year = value
    elif(header == 'Date'):
        metric.date =value
    elif(header == 'Base Price Elasticity'):
        metric.base_price_elasticity = round(float(value),3)
    elif(header == 'Cross Elasticity'):
        metric.cross_elasticity = round(float(value),3)
    elif(header == 'Net Elasticity'):
        metric.net_elasticity = round(float(value),3)
    elif(header == 'Base Units'):
        metric.base_units = round(float(value),3)
    elif(header == 'List Price'):
        metric.list_price = round(float(value),3)
    elif(header == 'Retailer Median Base Price'):
        metric.retailer_median_base_price = round(float(value),3)
    elif(header == 'Retailer Median Base Price  w\o VAT'):
        metric.retailer_median_base_price_w_o_vat = round(float(value),3)
    elif(header == 'On Inv. %'):
        metric.on_inv_percent = round(float(value) * 100,3) 
    elif(header == 'Off Inv. %'):
        metric.off_inv_percent = round(float(value) * 100,3) 
    elif(header == 'TPR %'):
        metric.tpr_percent = round(float(value) * 100,3) 
    elif(header == 'GMAC%, LSV'):
        metric.gmac_percent_lsv = round(float(value) * 100,3)
    elif(header == 'Product Group Weight (grams)'):
        metric.product_group_weight = round(float(value),3)

def _updateMetricDup(metric:model.ScenarioPlannerMetrics , value,header):
    # print(value , "Value " , type(value) , " TYPE VALUE")
    if(header == 'Category'):
        metric.category = value.strip()
    elif(header == 'Product Group'):
        metric.product_group = value.strip()
    elif(header == 'Retailer'):
        metric.retailer = value.strip()
    elif(header == 'Brand Filter'):
        metric.brand_filter = value.strip()
    elif(header == 'Brand Format Filter'):
        metric.brand_format_filter = value.strip()
    elif(header == 'Strategic Cell Filter'):
        metric.strategic_cell_filter = value.strip()
    elif(header == 'Year'):
        metric.year = value + 1
    elif(header == 'Date'):
        metric.date =value
    elif(header == 'Base Price Elasticity'):
        metric.base_price_elasticity = round(float(value),3)
    elif(header == 'Cross Elasticity'):
        metric.cross_elasticity = round(float(value),3)
    elif(header == 'Net Elasticity'):
        metric.net_elasticity = round(float(value),3)
    elif(header == 'Base Units'):
        metric.base_units = round(float(value),3) * 2
    elif(header == 'List Price'):
        metric.list_price = round(float(value),3)
    elif(header == 'Retailer Median Base Price'):
        metric.retailer_median_base_price = round(float(value),3)
    elif(header == 'Retailer Median Base Price  w\o VAT'):
        metric.retailer_median_base_price_w_o_vat = round(float(value),3)
    elif(header == 'On Inv. %'):
        metric.on_inv_percent = round(float(value) * 100,3) 
    elif(header == 'Off Inv. %'):
        metric.off_inv_percent = round(float(value) * 100,3) 
    elif(header == 'TPR %'):
        metric.tpr_percent = round(float(value) * 100,3) 
    elif(header == 'GMAC%, LSV'):
        metric.gmac_percent_lsv = round(float(value) * 100,3)
    elif(header == 'Product Group Weight (grams)'):
        metric.product_group_weight = round(float(value),3)

def _genObj(obj,value,header):
    # print(value , "Value " , type(value) , " TYPE VALUE")
     
    if(header == 'Category'):
        obj["category"]= value.strip()
    elif(header == 'Product Group'):
        obj["product_group"]= value.strip()
    elif(header == 'Retailer'):
        obj["retailer"]= value.strip()
    elif(header == 'Brand Filter'):
        obj["brand_filter"]= value.strip()
    elif(header == 'Brand Format Filter'):
        obj["brand_format_filter"]= value.strip()
    elif(header == 'Strategic Cell Filter'):
        obj["strategic_cell_filter"]= value.strip()
    elif(header == 'Year'):
        obj["year"]= value
    elif(header == 'Date'):
        obj["date"] =value
    elif(header == 'Base Price Elasticity'):
        obj["base_price_elasticity"]= round(float(value),3)
    elif(header == 'Cross Elasticity'):
        obj["cross_elasticity"]= round(float(value),3)
    elif(header == 'Net Elasticity'):
        obj["net_elasticity"]= round(float(value),3)
    elif(header == 'Base Units'):
        obj["base_units"]= round(float(value),3)
    elif(header == 'List Price'):
        obj["list_price"]= round(float(value),3)
    elif(header == 'Retailer Median Base Price'):
        obj["retailer_median_base_price"]= round(float(value),3)
    elif(header == 'Retailer Median Base Price  w\o VAT'):
        obj["retailer_median_base_price_w_o_vat"]= round(float(value),3)
    elif(header == 'On Inv. %'):
        obj["on_inv_percent"]= round(float(value) * 100,3) 
    elif(header == 'Off Inv. %'):
        obj["off_inv_percent"]= round(float(value) * 100,3) 
    elif(header == 'TPR %'):
        obj["tpr_percent"]= round(float(value) * 100,3) 
    elif(header == 'GMAC%, LSV'):
        obj["gmac_percent_lsv"]= round(float(value) * 100,3)
    elif(header == 'Product Group Weight (grams)'):
        obj["product_group_weight"]= round(float(value),3)
        
        
def lift(file1,file2):
    retailer = 'Tander'
    ppg = 'A.Korkunov 192g'
    segment = 'BOXES'
    base_value = roi.main(file1,file2,retailer , ppg , segment)
    # len(base_value)
    # base_value['Incrementa']
    # base_value['Base']
    # print(base_value , "base value")
    data = model.ModelData.objects.select_related('model_meta').filter(
        model_meta__account_name = retailer,
        model_meta__corporate_segment = segment,
        model_meta__product_group = ppg
    ).order_by('week')
    print(len(data) , "len of data")
    # import pdb
    # pdb.set_trace()
    for i in range(0,len(data)):
        query = data.get(week = i+1)
        # import pdb
        # pdb.set_trace()
        query.incremental_unit = decimal.Decimal(round(base_value['Incremental'][i],6))
        query.base_unit =  decimal.Decimal(round(base_value['Base'][i],6))
        query.save()
        print('saved')
    # import pdb
    # pdb.set_trace()
    

def lift_test():
    from django.db.models.query import Prefetch
    from django.forms.models import model_to_dict
    from . import constants as const
    import pandas as pd
    
    
    retailer = "Tander"
    ppg = 'A.Korkunov 192g'
    coeff_values = [ 'model_meta__id','model_meta__account_name', 
                    'model_meta__corporate_segment', 'model_meta__product_group', 'model_meta__brand_filter',
                    'model_meta__brand_format_filter', 'model_meta__strategic_cell_filter','wmape', 'rsq',
                    'intercept', 'median_base_price_log', 'tpr_discount', 'tpr_discount_lag1',
                    'tpr_discount_lag2', 'catalogue', 'display', 'acv', 'si', 
                    'si_month', 'si_quarter', 'c_1_crossretailer_discount', 'c_1_crossretailer_log_price', 'c_1_intra_discount', 
                    'c_2_intra_discount', 'c_3_intra_discount', 'c_4_intra_discount', 'c_5_intra_discount',
                    'c_1_intra_log_price', 'c_2_intra_log_price', 'c_3_intra_log_price', 'c_4_intra_log_price', 'c_5_intra_log_price', 'category_trend', 'trend_month', 'trend_quarter', 'trend_year', 'month_no', 'flag_promotype_motivation', 'flag_promotype_n_pls_1', 'flag_promotype_traffic', 'flag_nonpromo_1', 'flag_nonpromo_2', 'flag_nonpromo_3', 'flag_promo_1', 'flag_promo_2', 'flag_promo_3', 'holiday_flag_1', 'holiday_flag_2', 'holiday_flag_3', 'holiday_flag_4', 'holiday_flag_5', 'holiday_flag_6', 'holiday_flag_7', 'holiday_flag_8', 'holiday_flag_9', 'holiday_flag_10' 
                    ]
    data_values = ['model_meta__id','model_meta__account_name', 
                    'model_meta__corporate_segment', 'model_meta__product_group', 'model_meta__brand_filter',
                    'model_meta__brand_format_filter', 'model_meta__strategic_cell_filter',
                    'year','quater','month','period','date','week',
                    'intercept', 'median_base_price_log', 'tpr_discount', 'tpr_discount_lag1',
                    'tpr_discount_lag2', 'catalogue', 'display', 'acv', 'si', 
                    'si_month', 'si_quarter', 'c_1_crossretailer_discount', 'c_1_crossretailer_log_price', 'c_1_intra_discount', 
                    'c_2_intra_discount', 'c_3_intra_discount', 'c_4_intra_discount', 'c_5_intra_discount',
                    'c_1_intra_log_price', 'c_2_intra_log_price', 'c_3_intra_log_price', 'c_4_intra_log_price', 'c_5_intra_log_price', 'category_trend', 'trend_month', 'trend_quarter', 'trend_year', 'month_no', 'flag_promotype_motivation', 'flag_promotype_n_pls_1', 'flag_promotype_traffic', 'flag_nonpromo_1', 'flag_nonpromo_2', 'flag_nonpromo_3', 'flag_promo_1', 'flag_promo_2', 'flag_promo_3', 'holiday_flag_1', 'holiday_flag_2', 'holiday_flag_3', 'holiday_flag_4', 'holiday_flag_5', 'holiday_flag_6', 'holiday_flag_7', 'holiday_flag_8', 'holiday_flag_9', 'holiday_flag_10', 
                    'wk_sold_avg_price_byppg',
                    'average_weight_in_grams','weighted_weight_in_grams']
    coefficient = model.ModelCoefficient.objects.select_related('model_meta').filter(
        model_meta__account_name = retailer,
        model_meta__product_group = ppg
    ).values_list(*coeff_values)
    data = model.ModelData.objects.select_related('model_meta').filter(
        model_meta__account_name = retailer,
        model_meta__product_group = ppg
    ).values_list(*data_values)
    coeff_list = [list(i) for i in coefficient]
    data_list = [list(i) for i in data]
    print("-------------------------------------coeff-------------------------------------------------------")
    print()
    print()
    print(coeff_list ,"coeff_list LIST")
    print("----------------------------------------data----------------------------------------------------")
    print()
    print()
    print(data_list , "data_list LIST")
    print("--------------------------------------------------------------------------------------------")
    # pd.DataFrame
    # roi = model.ModelROI.objects.select_related('model_meta').filter(
    #     model_meta__account_name = retailer,
    #     model_meta__product_group = ppg
    # ).values_list()
    # import pdb
    # pdb.set_trace()
    
    # query  = queryset.filter(account_name = retailer , product_group = ppg)
    # coeff = [list(coeff) for coeff in query[0].prefetched_data]
    # query.values_list('account_name','corporate_segment','product_group','brand_filter')
    # print(query , "query")
    # print(coeff , "coeff")
    # print(list(query) , "list query")
    # import pdb
    # pdb.set_trace()
    # meta = model_to_dict(query)
    # import pdb
    # pdb.set_trace()
    # roi = [{**model_to_dict(d),**meta} for d in query.prefetched_roi]
    # data = [model_to_dict(d) for d in query.prefetched_data]
    # coeff = [{**model_to_dict(d),**meta} for d in query.prefetched_coeff]
    # print(meta , "meta")
    # print(roi , "droi")
    # print(data , "data")
    # print(coeff , "coeff")
    # print({**const.PROMO_MODEL_COEFF_MAP,**const.PROMO_MODEL_META_MAP} , "PROMO_MODEL_COEFF_MAP")
    # print( , "promo model meta map")
