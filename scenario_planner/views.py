# from xlwt import Workbook
from os import stat
import re
from django.core.exceptions import ObjectDoesNotExist
from utils import util
from rest_framework import fields, viewsets, mixins
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action, api_view
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import JSONRenderer
from rest_framework import status
from django.db.models.query import Prefetch
from django.shortcuts import get_object_or_404
from core import models as model
from core.models import ModelMeta, ModelROI, Scenario , ScenarioPlannerMetrics,ModelData,ModelCoefficient
from scenario_planner import serializers as sc
from rest_framework import serializers
 
from . import mixins as mixin

from utils import exceptions as exception
from utils import excel as excel
from utils import optimizer as optimizer
from json import loads, dumps
import ast
# import xlwt
# from xlrd import ope
from xlwt import Workbook
from django.http import HttpResponse
# import StringIO
import xlsxwriter
import json
import io
import itertools
import decimal
import math
import openpyxl
import pandas as pd
from utils import constants as CONST

def savePromo():
    return  {'account_name': 'Lenta', 'corporate_segment': 'BOXES', 'strategic_cell': 'cell', 'brand': 'brand', 'brand_format': 'format', 
'product_group': 'A.Korkunov 192g', 'promo_elasticity': 0, 'param_depth_all': 0, 
'week-1': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-2': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-3': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-4': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-5': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-6': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-7': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-8': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-9': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 
'week-10': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-11': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-12': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-13': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-14': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-15': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-16': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-17': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-18': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-19': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-20': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-21': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-22': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-23': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-24': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-25': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-26': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-27': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-28': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-29': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-30': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-31': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-32': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-33': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-34': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-35': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-36': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-37': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-38': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-39': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-40': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-41': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-42': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-43': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-44': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-45': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-46': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-47': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-48': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-49': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-50': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}, 'week-51': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0},
'week-52': {'promo_depth': 0, 'promo_mechanics': None, 'co_investment': 0}}


class MapPricingPromo(APIView):
    serializer_class = sc.MapPricingPromoSerializer
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    # def get(self, request, format=None):
    #     content = optimizer.process()
    #     return Response(content)
    
    def create_promo(self,request):
        # import pdb
        # pdb.set_trace()
        value =  ast.literal_eval(request.data['promo_details'].strip())
        ps = model.PricingSave.objects.get(id=int(request.data['pricing_id']))
        pr_save = model.PromoSave(
            account_name = ps.account_name,
            corporate_segment = ps.corporate_segment,
            product_group = ps.product_group,
            saved_scenario = ps.saved_scenario,
            saved_pricing = ps,
            promo_elasticity = value['promo_elasticity']
            
        )
        # value = savePromo()
        # scenario = model.SavedScenario(
        #     scenario_type = 'promo',
        #     name =  'promo1',
        #     comments = 'promo1',
        #     user = request.user
            
        # )   
        # scenario.save()
        # pr_save = model.PromoSave(
        #         account_name = value['account_name'],
        #         corporate_segment = value['corporate_segment'],
        #         product_group =value['product_group'],
        #         promo_elasticity = value['promo_elasticity'],
        #         saved_scenario = scenario
        #     )
        pr_save.save()
        bulk_pricing_week = []
        for i in value.keys():
            week_regex = util._regex(r'week-\d{1,2}',i)
            if week_regex:
                week = int(util._regex(r'\d{1,2}',week_regex.group()).group())
                # print(week , "week")
                # print(value[i] , "week value")
                pw = model.PromoWeek(
                   
                    week = week,
                    year = 2021,
                    promo_depth = value[i]['promo_depth'],
                     co_investment = value[i]['co_investment'],
                      promo_mechanic =value[i]['promo_mechanics'],
                       pricing_save = pr_save,
                )
                bulk_pricing_week.append(pw)
        model.PromoWeek.objects.bulk_create(bulk_pricing_week)        

    
    def post(self, request, format=None):
        self.create_promo(request)
        # import pdb
        # pdb.set_trace()
        
            
        return Response({}, status=status.HTTP_201_CREATED)
    

class SaveScenarioViewSet(viewsets.ModelViewSet):
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = model.SavedScenario.objects.all()
    serializer_class = sc.SaveScenarioSerializer
    serializer_classes = {
        'promo' : sc.SaveScenarioSerializer
    }
    def dispatch(self, request, *args, **kwargs):
        print(request , "request from dispatch")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        print(self.request , "request from get query set")
        return super().get_queryset()
    def get_serializer_class(self):
        print(self.request , "request from serializer")
        return super().get_serializer_class()
    
    def list(self, request, *args, **kwargs):
        print(request , "request from list")
        
        return super().list(request, *args, **kwargs)
    def create_promo(self,request):
        value = savePromo()
        scenario = model.SavedScenario(
            scenario_type = 'promo',
            name =  'promo1',
            comments = 'promo1',
            user = request.user
            
        )   
        scenario.save()
        pr_save = model.PromoSave(
                account_name = value['account_name'],
                corporate_segment = value['corporate_segment'],
                product_group =value['product_group'],
                promo_elasticity = value['promo_elasticity'],
                saved_scenario = scenario
            )
        pr_save.save()
        bulk_pricing_week = []
        for i in value.keys():
            week_regex = util._regex(r'week-\d{1,2}',i)
            if week_regex:
                week = int(util._regex(r'\d{1,2}',week_regex.group()).group())
                # print(week , "week")
                # print(value[i] , "week value")
                pw = model.PromoWeek(
                   
                    week = week,
                    year = 2021,
                    promo_depth = value[i]['promo_depth'],
                     co_investment = value[i]['co_investment'],
                      promo_mechanic =value[i]['promo_mechanics'],
                       pricing_save = pr_save,
                )
                bulk_pricing_week.append(pw)
        model.PromoWeek.objects.bulk_create(bulk_pricing_week)        
    
    
    def create(self, request, *args, **kwargs):
        # import pdb
        # pdb.set_trace()
        if 'scenario_type' in request.data:
            if request.data['scenario_type'] == 'promo':
                self.create_promo(request)
                return Response({}, 200)
        request.data['name']
        request.data['comments']
        request_dump = json.loads(request.data['savedump'])
        price_change =  request_dump['formArray']
        product = request_dump['productFilter']
        retailer = request_dump['retailerFilter']
        p_r_list = list(itertools.product([i.replace('Magnit','Tander') for i in retailer],product)) # change backend data in pricing scenario
        model_meta_set = set(model.ModelMeta.objects.values_list('account_name','product_group'))
        model_meta_list = []
        for i in model_meta_set:
            r = util.remove_duplicate_spaces(i[0]),util.remove_duplicate_spaces(i[1])
            model_meta_list.append(r)
        available = []
        for i in p_r_list:
            if (util.remove_duplicate_spaces(i[0]),util.remove_duplicate_spaces(i[1])) in model_meta_list:
                available.append(i)
        # import pdb
        # pdb.set_trace()
        scenario = model.SavedScenario(
            scenario_type = 'pricing',
            name =  request.data['name'],
            comments = request.data['comments'],
            user = self.request.user
            
        )   
        scenario.save()
        bulk_pricing_week = []
        for i in available:
            pr_save = model.PricingSave(
                account_name = i[0],
                corporate_segment = i[0],
                product_group = i[1],
                saved_scenario = scenario
            )
            pr_save.save()
            
            sc = ScenarioPlannerMetrics.objects.filter(retailer = i[0].replace('Tander','Magnit') , product_group = i[1])
            price = [d for d in price_change if d['product_group'] == i[1]][0]
            # print(price.lpi_increase , "price information")
            # import pdb
            # pdb.set_trace()
            
            for sce in sc:
                cogs = sce.list_price * decimal.Decimal(1- abs(1*(sce.gmac_percent_lsv/100)))
                pw = model.PricingWeek(
                   
                    week = sce.week,
                    year = sce.year,
                    lp_increase = sce.list_price * decimal.Decimal(1 + int(price['lpi_increase'])/100),
                    rsp_increase = sce.retailer_median_base_price_w_o_vat * decimal.Decimal(1 + int(price['rsp_increase'])/100),
                    cogs_increase = cogs * decimal.Decimal(1 + int(price['cogs_increase'])/100),
                    pricing_save = pr_save,
                )
                bulk_pricing_week.append(pw)
        model.PricingWeek.objects.bulk_create(bulk_pricing_week)        
        
        # import pdb
        # pdb.set_trace()
        # scenario = model.SavedScenario(
        #     scenario_type = 'pricing',
        #     name =  request.data['name'],
        #     comments = request.data['comments'],
        #     user = self.request.user
            
        # )   
        # scenario.save()
        
        
        return Response({},status=200)
    
class ScenarioViewSet(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
mixins.UpdateModelMixin,mixins.DestroyModelMixin , mixins.RetrieveModelMixin):
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = Scenario.objects.all()
    serializer_class = sc.ScenarioSerializer
    lookup_field = "id"
    # serializer_ac

    def get_queryset(self):
        # queryset = super(ScenarioViewSet, self).get_queryset()
        return self.queryset.filter(user=self.request.user).order_by('-name')
    def filter_queryset(self,queryset):
        # print(self.request.GET.get('yearly' , '') , "yearly value")
        # print(type(self.request.GET.get('yearly' , '')) , "yearly type")
        if(self.request.GET.get('yearly' , '')!="ALL"):
            return queryset.filter(is_yearly = self.request.GET.get('yearly' , '') == 'true')
        return queryset
    
    
    def create(self, request , *args , **kwargs):
        serializer = self.get_serializer(data = self.request.data)
            # import pdb
            # pdb.set_trace()
        if serializer.is_valid():
            serializer.save(user = self.request.user)
        else:
           return Response(serializer.errors , status=status.HTTP_400_BAD_REQUEST)
        # if not request.data['name']:
        #     raise exception.EmptyException
        # query = self.queryset.filter(user=self.request.user)
        # if(query.count() > 20):
        #     raise exception.CountExceedException
        # if(query.filter(name = serializer.validated_data['name']).exists()):
        #     raise exception.AlredyExistsException
        # super().create(request,args,kwargs)
        return Response(serializer.data)
        

   
    # def perform_create(self, serializer):
    #     query = self.queryset.filter(user=self.request.user)
    #     if(query.count() > 2):
    #         raise exception.CountExceedException
    #     if(query.filter(name = serializer.validated_data['name']).exists()):
    #         raise exception.AlredyExistsException
    #     serializer.save(user=self.request.user)

    def put(self,**kwargs):
        # print(self.request , "request")
        # print(kwargs , "kwargs")
        self.partial_update(self.request,**kwargs)
    
    def delete(self, request, *args, **kwargs):
        return self.destroy(request, *args, **kwargs)

    # def
class ScenarioPlannerMetricsViewSet(viewsets.GenericViewSet, mixins.ListModelMixin):
    queryset = ScenarioPlannerMetrics.objects.all()
    serializer_class = sc.ScenarioPlannerMetricsSerializer

class ScenarioPlannerMetricsViewSetObject(viewsets.GenericViewSet, mixins.ListModelMixin):
    queryset = ScenarioPlannerMetrics.objects.all()
    serializer_class = sc.ScenarioPlannerMetricsSerializerObject
    
class LoadScenario(viewsets.ReadOnlyModelViewSet,mixin.CalculationMixin):
    # queryset = Scenario.objects.filter(scenario_type = 'promo')
    queryset = model.SavedScenario.objects.all()
    # serializer_class = sc.PromoScenarioSavedList
    serializer_class = sc.ScenarioSavedList
    lookup_field = "id"
    
    
    def retrieve_pricing_promo(self, request, *args, **kwargs):
        
        # import pdb
        # pdb.set_trace()
        # get_serializer = sc.ModelMetaGetSerializer()
        pricing_save_id = request.parser_context['kwargs']['_id']  # pricing save id
        pricing_week = model.PricingWeek.objects.select_related('pricing_save').filter(pricing_save__id = pricing_save_id )
        # import pdb
        # pdb.set_trace()
        # self.calculate_finacial_metrics_from_pricing(pricing_week)
        return Response(self.calculate_finacial_metrics_from_pricing(pricing_week), status = 200)
    def retrieve_bkp(self, request, *args, **kwargs):
        self.get_object()
        
        weeks = model.PromoWeek.objects.select_related('pricing_save','pricing_save__saved_scenario').filter(
         pricing_save__saved_scenario = self.get_object()
        )
        self.calculate_finacial_metrics(weeks)
        
        value_dict = ast.literal_eval(self.get_object().savedump)
        return Response( self.calculate_finacial_metrics(value_dict),
                        status=status.HTTP_201_CREATED)
        
    def retrieve(self, request, *args, **kwargs):
        obj = self.get_object()
        # import pdb
        # pdb.set_trace()
        if obj.scenario_type == 'pricing':
            pricing_save = model.PricingSave.objects.filter(saved_scenario = obj)
            # import pdb
            # pdb.set_trace()
            serializer = sc.PricingSaveSerializer(pricing_save,many=True)
            return Response(serializer.data , status=200)
    
        weeks = model.PromoWeek.objects.select_related('pricing_save','pricing_save__saved_scenario').filter(
            pricing_save__saved_scenario = obj
        )
       
        # for i in weeks:
        #     print(i)
        # print( weeks[0].pricing_save.account_name , 'accname')
        # print( weeks[0].pricing_save.product_group , 'product group')
        # print( weeks[0].pricing_save.promo_elasticity , 'promo_elasticity')
            
        
        
        # value_dict = ast.literal_eval(self.get_object().savedump)
        return Response( self.calculate_finacial_metrics(weeks),
                        
                    status=status.HTTP_201_CREATED)
        # return Response( {},
                        
        #             status=status.HTTP_201_CREATED)

        
    # @action(methods=['get'], detail=True)
    # def detail(self, *args, **kwargs):
    #     return Response({
    #         "rr"  : "dddd"
    #     })

class ExampleViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Scenario.objects.all()
    serializer_class = sc.ScenarioSerializer


    @action(methods=['post'], detail=True)
    def download(self, *args, **kwargs):

        formdata = self.request.data['data']
        typ = self.request.data['type']
       
        output = io.BytesIO()
        if typ == 'comp':
            excel.excel_summary(json.loads(formdata) , output)
        else:
            excel.excel(json.loads(formdata) , output)
        output.seek(0)
        filename = 'django_simple.xlsx'
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

    @action(methods=['post'], detail=True)
    def getData(self, *args, **kwargs):
        # print(args , "ARGS")
        # print(kwargs , "KWARGS")
        formdata = self.request.data['file']
        # print(formdata)
        # import pdb
        # pdb.set_trace()
        # print(self , "self")
        return 1
       
        # instance = self.get_object()

        # # get an open file handle (I'm just using a file attached to the model for this example):
        # file_handle = instance.file.open()

        # # send file
        # response = FileResponse(file_handle, content_type='whatever')
        # response['Content-Length'] = instance.file.size
        # response['Content-Disposition'] = 'attachment; filename="%s"' % instance.file.name

        # return response

def down(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    expenses = (
        ['Rent', 1000],
        ['Gas',   100],
        ['Food',  300],
        ['Gym',    50],
    )

    row = 3
    col = 3

    # Iterate over the data and write it out row by row.
    for item, cost in (expenses):
        worksheet.write(row, col,     item)
        worksheet.write(row, col + 1, cost)
        row += 1

    worksheet.write(row, 3, 'Total')
    worksheet.write(row, 4, '=SUM(B1:B4)')

    workbook.close()
    output.seek(0)
    filename = 'django_simple.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    return response


def WriteToExcel(weather_data, town=None):
    # output = StringIO.StringIO()
    workbook = xlsxwriter.Workbook("output")

    # Here we will adding the code to add data

    workbook.close()
    xlsx_data = " output.getvalue()"
    # xlsx_data contains the Excel file
    return xlsx_data

class ModelOptimize(APIView):
    serializer_class = sc.CommentSerializer
    def get(self, request, format=None):
        content = optimizer.process()
        return Response(content)
    
    def post(self, request, format=None):
        content = None
        
        serializer = sc.CommentSerializer(data=request.data)
        
        if serializer.is_valid():
            
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        return Response(content, status=status.HTTP_201_CREATED)

# class PromoSimulatorView(viewsets.GenericViewSet,mixins.ListModelMixin):
#     queryset = ModelData.objects.select_related('model_meta').prefetch_related(
#         Prefetch('model_meta__coefficient',
#                  queryset=ModelCoefficient.objects.all(),
#                  to_attr='prefetched_coeff')
#         ).order_by('id')[:10]
     
#     serializer_class = sc.ModelDataSerializer
#     # serializer_class = sc.PromoSimulatorSerializer
#     def get(self, request, format=None):
#         serializer = sc.ModelDataSerializer()
#         return Response(serializer.data, status=status.HTTP_201_CREATED)
#     def post(self, request, format=None):
#         serializer = sc.ModelDataSerializer()
#         return Response(serializer.data, status=status.HTTP_201_CREATED)

class PromoSimulatorView(viewsets.GenericViewSet,mixin.CalculationMixin):
    queryset = ModelMeta.objects.prefetch_related(
        Prefetch(
            'data',
        queryset = ModelData.objects.all().order_by('week'),
        to_attr='prefetched_data'
        ),
        # 'data',
        Prefetch(
            'coefficient',
            queryset=ModelCoefficient.objects.all(),
            to_attr='prefetched_coeff'
        ),
        Prefetch(
            'roi',
            queryset=ModelROI.objects.all().order_by('week'),
            to_attr='prefetched_roi'
        )
    ).order_by('id')
     
    def get(self, request, format=None):
        
        serializer = sc.ModelMetaGetSerializer()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def get_serializer_class(self):
        return sc.ModelMetaGetSerializer
    
    def get_queryset(self):
        return super().get_queryset()
    
    def post(self, request, format=None):
        get_serializer = sc.ModelMetaGetSerializer(request.data)
        value_dict = loads(dumps((get_serializer.to_internal_value(request.data))))
        try:
            response = self.calculate_finacial_metrics_from_request(value_dict)
            return Response(response ,
                        status=status.HTTP_201_CREATED)
        except ObjectDoesNotExist as e:
            return Response({'error' : str(e)},status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error' : "Something went wrong!"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PromoSimulatorUploadView(viewsets.GenericViewSet,mixin.CalculationMixin):
    def get(self, request, format=None):
        serializer = sc.ModelMetaExcelUpload()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def get_serializer_class(self):
        return sc.ModelMetaExcelUpload
    
    def get_queryset(self):
        return super().get_queryset()

    def get_promo_mechanics(self,value):
        if value == "Motivation":
            return "Flag_promotype_Motivation"
        elif value == "N Pls 1":
            return "Flag_promotype_N_pls_1"
        elif value == "Traffic":
            return "Flag_promotype_traffic"
        elif value == "Promo depth":
            return "Flag_promotype_traffic"
        else:
            return None
    
    def post(self, request, format=None):
        get_serializer = sc.ModelMetaExcelUpload(request.data)
        csv_file = request.FILES["simulator_input"]
        workbook = openpyxl.load_workbook(csv_file,data_only=True)
        sheet_name = workbook.sheetnames
        sheet = workbook[sheet_name[0]]
        excel_data = sheet.values
        # Get the first line in file as a header line
        columns = next(excel_data)[0:]
        excel_input_df = pd.DataFrame(excel_data,columns=columns)

        # Validating the excel input 
        if excel_input_df.shape[0] != 52:
            return Response({'error' : "Please upload a excel with all 52 week values"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        if list(columns) != CONST.PROMO_SIMULATOR_EXCEL_INPUT_COLS:
            return Response({'error' : "Headers doesn't match, please upload a valid excel"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        account_name = excel_input_df['Account name'].unique()
        corporate_segment = excel_input_df['Corporate segment'].unique()
        strategic_cell = excel_input_df['Strategic cell'].unique()
        brand = excel_input_df['Brand'].unique()
        brand_format = excel_input_df['Brand format'].unique()
        product_group = excel_input_df['Product group'].unique()
        promo_elasticity = excel_input_df['Promo elasticity'].unique()

        if len(account_name) != 1:
            return Response({'error' : "Account Name should be unique"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        if len(corporate_segment) != 1:
            return Response({'error' : "Corporate segment should be unique"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        if len(strategic_cell) != 1:
            return Response({'error' : "Strategic cell should be unique"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        if len(brand) != 1:
            return Response({'error' : "Brand should be unique"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        if len(brand_format) != 1:
            return Response({'error' : "Brand format should be unique"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        if len(product_group) != 1:
            return Response({'error' : "Product group should be unique"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        if len(promo_elasticity) != 1:
            return Response({'error' : "Promo elasticity should be unique"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        unique_ppg_combination = excel_input_df["Account name"] + '-' + excel_input_df["Product group"]
        unique_ppg_combination = unique_ppg_combination.unique()
        if len(unique_ppg_combination) != 1:
            return Response({'error' : "Retailer and PPG Combination should be unique"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        co_investment = excel_input_df['Co investment'].between(0,100,inclusive = True)
        promo_depth = excel_input_df['Promo depth'].between(0,100,inclusive = True)
        if promo_depth.nunique() != 1:
            return Response({'error' : "Promo depth values should be in between 0 to 100"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        if co_investment.nunique() != 1:
            return Response({'error' : "Co-Investment values should be in between 0 to 100"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        excel_input_df['Promo elasticity'] = excel_input_df['Promo elasticity'].astype(int)
        excel_input_df['Promo depth'] = excel_input_df['Promo depth'].astype(float)
        excel_input_df['Co investment'] = excel_input_df['Co investment'].astype(float)


        simulator_input = {
            'account_name': account_name[0],
            'corporate_segment': corporate_segment[0],
            'strategic_cell': strategic_cell[0],
            'brand': brand[0],
            'brand_format': brand_format[0],
            'product_group': product_group[0],
            'promo_elasticity': promo_elasticity[0],
            'param_depth_all': 0,
        }
        for i in range(0,52):
            simulator_input["week-"+str(i+1)] = {
                'promo_depth': excel_input_df['Promo depth'][i], 
                'promo_mechanics': self.get_promo_mechanics(excel_input_df['Promo mechanics'][i]), 
                'co_investment': excel_input_df['Co investment'][i]
            }
        try:
            response = self.calculate_finacial_metrics_from_request(simulator_input)
            return Response(response ,
                        status=status.HTTP_201_CREATED)
        except ObjectDoesNotExist as e:
            return Response({'error' : str(e)},status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error' : "Something went wrong!"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            
        
class LoadScenarioTest(viewsets.GenericViewSet,mixin.CalculationMixin):
    
    queryset = model.SavedScenario.objects.all()
    serializer_class = sc.ScenarioSavedList
    lookup_field = "id"
    def get(self, request, format=None):
    
        serializer = sc.SavedScenario()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def get_serializer_class(self):
        return sc.SavedScenario

    def get_queryset(self):
        return super().get_queryset()

    def post(self, request, format=None):
        get_serializer = sc.ModelMetaGetSerializer(request.data)
        value_dict = loads(dumps((get_serializer.to_internal_value(request.data))))
        try:
            response = self.calculate_finacial_metrics_from_request(value_dict)
            return Response(response ,
                        status=status.HTTP_201_CREATED)
        except ObjectDoesNotExist as e:
            return Response({'error' : str(e)},status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error' : "Something went wrong!"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            
        
        
    
class PromoSimulatorTestViewSet(viewsets.ModelViewSet):
    queryset = ModelMeta.objects.all()
    serializers = sc.ModelMetaGetSerializerTest()
    # .prefetch_related(
        # 'data',
        # Prefetch(
        #     'coefficient',
        #     queryset=ModelCoefficient.objects.all(),
        #     to_attr='prefetched_coeff'
        # )
    # ).order_by('id')
     
    # def get(self, request, format=None):
    #     # query=ModelMeta.objects.all()
        
    #     serializer = sc.ModelMetaGetSerializerTest()
    #     return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def get_serializer_class(self):
        return sc.ModelMetaGetSerializerTest
    
    # def get_queryset(self):
    #     return super().get_queryset()
    
    def post(self, request, format=None):
        query = self.queryset
        get_serializer = sc.ModelMetaGetSerializerTest(request.data)
        value_dict = loads(dumps((get_serializer.to_internal_value(request.data))))
        # print(value_dict , "value dict")
        query = query.get(account_name = value_dict['account_name'],
                  corporate_segment=value_dict['corporate_segment'],
                  product_group = value_dict['product_group'])
        serializer = sc.ModelMetaSerializer(query)
        return Response(serializer.data, status=status.HTTP_201_CREATED)